from fastapi import FastAPI, HTTPException
from fastapi.responses import HTMLResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
import openpyxl
import json
import os
import re
import sys
from pathlib import Path
from openai import OpenAI

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

# ── LLM setup (same pattern as autoparts tool) ────────────────────────
_HUBGPT_KEY = os.environ.get("HUBGPT_API_KEY", "")
_SF_KEY     = os.environ.get("SF_API_KEY", "")

if _HUBGPT_KEY:
    _client   = OpenAI(api_key=_HUBGPT_KEY, base_url=os.environ.get("HUBGPT_BASE_URL", "https://hubgpt.corp.ebay.com/v1"))
    _MODEL    = os.environ.get("LLM_MODEL", "claude-haiku-4-5")
    _BACKEND  = "HubGPT"
elif _SF_KEY:
    _client   = OpenAI(api_key=_SF_KEY, base_url="https://api.siliconflow.cn/v1")
    _MODEL    = os.environ.get("LLM_MODEL", "deepseek-ai/DeepSeek-V3")
    _BACKEND  = "SiliconFlow"
else:
    _client   = None
    _MODEL    = ""
    _BACKEND  = "none"

print(f"LLM backend: {_BACKEND} / model: {_MODEL}")

app = FastAPI()
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

BASE_DIR = Path(__file__).parent
EXCEL_PATH = Path(r"C:\Users\jiancui\Desktop\表名理解\所有样本表表头汇总.xlsx")

CIRCLED = "①②③④⑤⑥⑦⑧⑨⑩⑪⑫⑬⑭⑮⑯⑰⑱"

TABLE_COLORS = {
    "dw_checkout_trans": "#1a6eb5",
    "dw_lstg_item":      "#2e7d32",
    "dw_cal_dt":         "#6a1b9a",
    "Cate_ENG":          "#e65100",
    "LEAF_ENG":          "#e65100",
    "CBT_SLR_vertical":  "#00695c",
    "Category_analysis": "#00695c",
    "Top_Seller_analysis":"#00695c",
    "AMS_ERS_NAME":      "#c62828",
    "SLR_BYR_MAPPING":   "#c62828",
}


def parse_excel():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb.active

    # Handle merged cells
    merged_top = {}
    for rng in ws.merged_cells.ranges:
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                if not (r == rng.min_row and c == rng.min_col):
                    merged_top[(r, c)] = (rng.min_row, rng.min_col)

    def cell(row, col):
        key = (row, col)
        if key in merged_top:
            r0, c0 = merged_top[key]
            return ws.cell(row=r0, column=c0).value
        return ws.cell(row=row, column=col).value

    # ── Pass 1: parse synonym groups from col E rows 2-57 ──────────────
    groups = {}          # group_name → {name, entries, verification}
    cur = None

    for rn in range(2, 58):
        e = ws.cell(row=rn, column=5).value
        if not e:
            continue
        e = str(e).strip()

        if e and e[0] in CIRCLED:
            f_raw = cell(rn, 6)
            ver = str(f_raw).strip() if f_raw else "— 未验证"
            cur = e
            groups[cur] = {"name": e, "entries": [], "verification": ver}
        elif cur and "→" in e:
            left, _, right = e.partition("→")
            left = left.strip()
            # split multiple fields by /
            raw_fields = [x.strip() for x in left.split("/") if x.strip()]
            cleaned = []
            for f in raw_fields:
                f = re.split(r"[（(]", f)[0].strip()
                if len(f) > 1:
                    cleaned.append(f)
            groups[cur]["entries"].append({
                "fields": cleaned,
                "tables": right.strip(),
                "raw": e.strip()
            })

    # ── Pass 2: field_name → group mapping ─────────────────────────────
    field_to_group = {}   # lower(field) → group_name
    group_synonyms = {}   # group_name → [{field, tables}]

    for gname, gdata in groups.items():
        syns = []
        for entry in gdata["entries"]:
            for f in entry["fields"]:
                field_to_group[f.lower()] = gname
                syns.append({"field": f, "tables": entry["tables"]})
        group_synonyms[gname] = syns

    # ── Pass 3: parse main field rows ──────────────────────────────────
    fields = []
    for rn in range(2, ws.max_row + 1):
        t_raw = ws.cell(row=rn, column=1).value
        fn    = ws.cell(row=rn, column=3).value
        desc  = ws.cell(row=rn, column=4).value
        if not fn:
            continue

        t_raw = str(t_raw).strip() if t_raw else ""
        parts = t_raw.split(" ", 1)
        table_en = parts[0]
        table_cn = parts[1] if len(parts) > 1 else ""

        fn = str(fn).strip()
        gname = field_to_group.get(fn.lower())
        synonyms = []
        if gname:
            synonyms = [s for s in group_synonyms[gname] if s["field"].lower() != fn.lower()]

        fields.append({
            "table":        table_en,
            "table_cn":     table_cn,
            "table_color":  TABLE_COLORS.get(table_en, "#546e7a"),
            "field":        fn,
            "description":  str(desc).strip() if desc else "",
            "group":        gname or "",
            "verification": groups[gname]["verification"] if gname else "",
            "synonyms":     synonyms,
        })

    wb.close()
    return {"fields": fields, "groups": list(groups.values())}


print("Loading Excel …")
DATA = parse_excel()
print(f"Loaded {len(DATA['fields'])} fields, {len(DATA['groups'])} synonym groups")


@app.get("/", response_class=HTMLResponse)
async def root():
    return (BASE_DIR / "index.html").read_text(encoding="utf-8")


@app.get("/api/fields")
async def get_fields():
    return DATA


class AIReq(BaseModel):
    query: str
    context: str
    mode: str = "search"   # "search" | "generate"


@app.get("/api/backend")
async def backend_info():
    return {"backend": _BACKEND, "model": _MODEL, "ready": _client is not None}


@app.post("/api/ai")
async def ai_query(req: AIReq):
    if _client is None:
        raise HTTPException(status_code=503, detail="未配置 LLM。请设置 HUBGPT_API_KEY 或 SF_API_KEY 环境变量后重启。")

    if req.mode == "search":
        system = (
            "你是eBay数据仓库字段专家。用户搜索了某个字段/指标但没找到精确匹配。"
            "请从提供的字段列表中找出最相关的字段（最多6个），用中文解释相关原因。\n"
            '返回JSON：{"suggestions":[{"field":"字段名","table":"表名","description":"字段含义","reason":"相关原因"}]}'
        )
        user_msg = f"用户查询：{req.query}\n\n字段列表（table | field | description）：\n{req.context}"
    else:
        system = (
            "你是eBay数据仓库专家，同时熟悉Tableau Desktop。"
            "用户需要一个现有字段中没有的计算指标。请生成：\n"
            "1. Spark SQL计算字段（可直接放在SELECT中的表达式）\n"
            "2. Tableau计算字段公式\n"
            '返回JSON：{"field_name":"建议字段名(英文下划线)","sql":"SQL公式","tableau":"Tableau公式","note":"简短说明"}'
        )
        user_msg = f"需要的指标：{req.query}\n\n可参考的相关字段：\n{req.context}"

    try:
        resp = _client.chat.completions.create(
            model=_MODEL,
            max_tokens=1024,
            messages=[
                {"role": "system", "content": system},
                {"role": "user",   "content": user_msg},
            ],
            timeout=60,
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

    text = resp.choices[0].message.content or ""
    m = re.search(r"\{.*\}", text, re.DOTALL)
    if m:
        try:
            return json.loads(m.group())
        except Exception:
            pass
    return {"raw": text}


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="127.0.0.1", port=8082)
